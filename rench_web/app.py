import os
import psycopg2
import psycopg2.extras
import hashlib
import difflib
import unicodedata
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, g, session
from functools import wraps
from urllib.parse import urlparse

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'rench_estoque_2026_segredo')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_PADRAO = os.path.join(os.path.dirname(BASE_DIR), 'data_atual.xlsx')

# Usar Supabase/Postgres via variavel de ambiente DATABASE_URL
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres.yfxfmwrasjukbsjjqbzs:kaio82046697@aws-1-us-west-2.pooler.supabase.com:6543/postgres')

PREFIXOS_RASTREIO = {
    "impressora": "IMP",
    "desktop": "CPU",
    "notebook": "NOT",
    "servidor": "SRV",
    "monitor": "MON",
    "periferico": "PER",
}

def normalizar_busca(texto):
    texto = str(texto or "").lower().strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return "".join(ch if ch.isalnum() else " " for ch in texto)

def preparar_termos_busca(texto):
    normalizado = normalizar_busca(texto)
    return [termo for termo in normalizado.split() if termo]

def texto_combinado(*valores):
    return normalizar_busca(" ".join(str(valor or "") for valor in valores))

def calcular_pontuacao_busca(termos, *valores):
    alvo = texto_combinado(*valores)
    if not termos or not alvo:
        return 0

    pontuacao = 0
    palavras = alvo.split()
    for termo in termos:
        if termo in alvo:
            pontuacao += 100
            continue

        melhor = 0
        for palavra in palavras:
            similaridade = difflib.SequenceMatcher(None, termo, palavra).ratio()
            if similaridade > melhor:
                melhor = similaridade

        if melhor >= 0.78:
            pontuacao += int(melhor * 80)
        elif melhor >= 0.68:
            pontuacao += int(melhor * 55)

    return pontuacao

def buscar_sugestoes_locais(cur, busca, limite=6):
    termos = preparar_termos_busca(busca)
    if not termos:
        return []

    cur.execute("""
        SELECT e.nome as empresa_nome, e.tipo as empresa_tipo, u.nome as unidade_nome, u.setor
        FROM empresas e
        LEFT JOIN unidades u ON u.empresa_id = e.id AND u.ativo=1
        WHERE e.ativo=1
    """)

    sugestoes = []
    for local in cur.fetchall():
        pontuacao = calcular_pontuacao_busca(
            termos,
            local["empresa_nome"],
            local["empresa_tipo"],
            local["unidade_nome"],
            local["setor"],
        )
        if pontuacao:
            sugestoes.append({
                "empresa": local["empresa_nome"],
                "tipo": local["empresa_tipo"],
                "unidade": local["unidade_nome"],
                "setor": local["setor"],
                "pontuacao": pontuacao,
            })

    sugestoes.sort(key=lambda item: item["pontuacao"], reverse=True)
    return sugestoes[:limite]

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        result = urlparse(DATABASE_URL)
        db = g._database = psycopg2.connect(
            host=result.hostname,
            port=result.port or 6543,
            user=result.username,
            password=result.password,
            dbname=result.path.lstrip('/'),
            connect_timeout=10
        )
        db.cursor_factory = psycopg2.extras.RealDictCursor
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def _executar_migrations():
    db = get_db()
    cur = db.cursor()

    def coluna_existe(tabela, coluna):
        cur.execute("""
            SELECT 1 FROM information_schema.columns
            WHERE table_schema='public' AND table_name=%s AND column_name=%s
        """, (tabela, coluna))
        return cur.fetchone() is not None

    migracoes = []
    for tabela, coluna, tipo in [
        ('unidades', 'setor', 'TEXT'),
        ('equipamentos', 'funcao', 'VARCHAR(50)'),
        ('equipamentos', 'tipo_impressao', 'VARCHAR(50)'),
        ('equipamentos', 'tamanho_papel', 'VARCHAR(50)'),
        ('equipamentos', 'funcionalidades', 'TEXT'),
        ('equipamentos', 'processador_modelo', 'VARCHAR(100)'),
        ('equipamentos', 'processador_geracao', 'VARCHAR(50)'),
        ('equipamentos', 'processador_velocidade', 'VARCHAR(50)'),
        ('equipamentos', 'memoria_capacidade', 'VARCHAR(50)'),
        ('equipamentos', 'memoria_tipo', 'VARCHAR(50)'),
        ('equipamentos', 'memoria_barramento', 'VARCHAR(50)'),
        ('equipamentos', 'memoria_velocidade', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_1_capacidade', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_1_tipo', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_2_capacidade', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_2_tipo', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_3_capacidade', 'VARCHAR(50)'),
        ('equipamentos', 'armazenamento_3_tipo', 'VARCHAR(50)'),
        ('equipamentos', 'local_atual_id', 'INTEGER'),
        ('equipamentos', 'setor_equipamento', 'VARCHAR(255)'),
        ('historico_defeitos', 'status', 'VARCHAR(50)'),
        ('suprimentos_itens', 'motivo_padrao', 'VARCHAR(50)'),
        ('suprimentos_itens', 'defeito', 'TEXT'),
        ('suprimentos_itens', 'estorno_estoque', 'INTEGER DEFAULT 0'),
    ]:
        if not coluna_existe(tabela, coluna):
            migracoes.append(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}")

    for sql in migracoes:
        try:
            cur.execute(sql)
        except Exception as e:
            print(f"Migration warning: {e}")
    db.commit()
    cur.close()

def init_db():
    db = get_db()
    cur = db.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS empresas (
            id SERIAL PRIMARY KEY,
            nome VARCHAR(255) NOT NULL,
            tipo VARCHAR(50) NOT NULL DEFAULT 'cliente',
            ativo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS unidades (
            id SERIAL PRIMARY KEY,
            empresa_id INTEGER NOT NULL REFERENCES empresas(id),
            nome VARCHAR(255) NOT NULL,
            setor TEXT,
            ativo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS equipamentos (
            id SERIAL PRIMARY KEY,
            codigo VARCHAR(100),
            tipo_equipamento VARCHAR(100),
            fabricante VARCHAR(255),
            modelo VARCHAR(255),
            numero_serie VARCHAR(255),
            patrimonio VARCHAR(255),
            status VARCHAR(100) DEFAULT 'ativo',
            unidade_id INTEGER REFERENCES unidades(id),
            local_atual_nome VARCHAR(255),
            cliente_atual VARCHAR(255),
            observacoes TEXT,
            contador_mono INTEGER DEFAULT 0,
            contador_color INTEGER DEFAULT 0,
            data_cadastro VARCHAR(50),
            ativo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS movimentacoes (
            id SERIAL PRIMARY KEY,
            equipamento_id INTEGER NOT NULL REFERENCES equipamentos(id),
            data_movimentacao VARCHAR(50) NOT NULL,
            origem_local VARCHAR(255),
            origem_unidade VARCHAR(255),
            destino_local VARCHAR(255) NOT NULL,
            destino_unidade VARCHAR(255),
            tipo_movimento VARCHAR(100) NOT NULL,
            responsavel VARCHAR(255),
            observacoes TEXT,
            contador_mono_anterior INTEGER,
            contador_mono_novo INTEGER,
            contador_color_anterior INTEGER,
            contador_color_novo INTEGER
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS historico_defeitos (
            id SERIAL PRIMARY KEY,
            equipamento_id INTEGER NOT NULL REFERENCES equipamentos(id),
            descricao TEXT,
            data_ocorrencia VARCHAR(50)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            nome VARCHAR(255) NOT NULL,
            usuario VARCHAR(255) UNIQUE NOT NULL,
            senha_hash VARCHAR(255) NOT NULL,
            nivel VARCHAR(50) DEFAULT 'tecnico',
            ativo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS acesso_remoto (
            id SERIAL PRIMARY KEY,
            local VARCHAR(255) NOT NULL,
            anydesk VARCHAR(255),
            senha VARCHAR(255),
            observacoes TEXT,
            ativo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS suprimentos_entregas (
            id SERIAL PRIMARY KEY,
            unidade_id INTEGER NOT NULL REFERENCES unidades(id),
            data_entrega VARCHAR(50) NOT NULL,
            responsavel VARCHAR(255),
            observacoes TEXT,
            data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS suprimentos_itens (
            id SERIAL PRIMARY KEY,
            entrega_id INTEGER NOT NULL REFERENCES suprimentos_entregas(id) ON DELETE CASCADE,
            tipo_suprimento VARCHAR(100) NOT NULL,
            modelo_impressora VARCHAR(255),
            quantidade INTEGER NOT NULL DEFAULT 1,
            motivo TEXT,
            motivo_padrao VARCHAR(50),
            defeito TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS estoque (
            id SERIAL PRIMARY KEY,
            tipo_suprimento VARCHAR(100) NOT NULL,
            modelo_impressora VARCHAR(255) NOT NULL DEFAULT '',
            quantidade INTEGER NOT NULL DEFAULT 0 CHECK (quantidade >= 0),
            estoque_minimo INTEGER NOT NULL DEFAULT 1,
            data_atualizacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_estoque_tipo_modelo
        ON estoque (tipo_suprimento, modelo_impressora)
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS estoque_movimentacoes (
            id SERIAL PRIMARY KEY,
            estoque_id INTEGER NOT NULL REFERENCES estoque(id) ON DELETE CASCADE,
            tipo_movimento VARCHAR(50) NOT NULL,
            quantidade INTEGER NOT NULL,
            saldo_antes INTEGER NOT NULL,
            saldo_depois INTEGER NOT NULL,
            motivo TEXT,
            responsavel VARCHAR(255),
            entrega_id INTEGER,
            data_movimento TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    db.commit()
    cur.close()

    _executar_migrations()

    db = get_db()
    cur = db.cursor()
    senha_hash = hashlib.sha256('ipascnma'.encode()).hexdigest()
    cur.execute("""
        INSERT INTO usuarios (nome, usuario, senha_hash, nivel)
        VALUES ('Administrador', 'admin', %s, 'admin')
        ON CONFLICT (usuario) DO NOTHING
    """, (senha_hash,))

    cur.execute("SELECT COUNT(*) as count FROM empresas")
    if cur.fetchone()['count'] == 0:
        padrao_empresas = [
            ('RENCH', 'rench'),
            ('SAMIR', 'cliente'),
            ('AES', 'cliente'),
            ('Mercadão', 'cliente'),
            ('Stuttgart Porsche', 'cliente'),
            ('R10', 'assistencia'),
        ]
        for nome, tipo in padrao_empresas:
            cur.execute("INSERT INTO empresas (nome, tipo) VALUES (%s, %s)", (nome, tipo))
        cur.execute("SELECT id FROM empresas WHERE nome='RENCH'")
        rench_id = cur.fetchone()['id']
        cur.execute("INSERT INTO unidades (empresa_id, nome, setor) VALUES (%s, %s, %s)", (rench_id, 'RENCH - Estoque', 'Estoque principal'))
        cur.execute("INSERT INTO unidades (empresa_id, nome, setor) VALUES (%s, %s, %s)", (rench_id, 'RENCH - Depósito', 'Depósito secundário'))
        cur.execute("SELECT id FROM empresas WHERE nome='R10'")
        r10_id = cur.fetchone()['id']
        cur.execute("INSERT INTO unidades (empresa_id, nome, setor) VALUES (%s, %s, %s)", (r10_id, 'R10 - Matriz SP', 'Assistência técnica'))

    db.commit()
    cur.close()
    print("Banco de dados inicializado!")

def _normalizar_texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto if texto else None

def _normalizar_data(valor):
    if valor is None:
        return None
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d")
    return str(valor)[:10]

def _mapear_tipo_movimento(valor):
    texto = (_normalizar_texto(valor) or "").lower()
    if "envio" in texto and "cliente" in texto:
        return "saida_cliente"
    if "retorno" in texto and "cliente" in texto:
        return "retorno_cliente"
    if "envio" in texto and ("manuten" in texto or "manutenc" in texto):
        return "envio_manutencao"
    if "retorno" in texto and ("manuten" in texto or "manutenc" in texto):
        return "retorno_manutencao"
    if "entrada" in texto:
        return "entrada_estoque"
    return "transferencia"

def _mapear_tipo_computador(valor):
    texto = (_normalizar_texto(valor) or "").lower()
    if "server" in texto or "servidor" in texto:
        return "servidor"
    if "note" in texto:
        return "notebook"
    if "monitor" in texto:
        return "monitor"
    if "desktop" in texto or "comput" in texto:
        return "desktop"
    return "periferico"

def _descobrir_fabricante(modelo):
    texto = (_normalizar_texto(modelo) or "").upper()
    marcas = ["OKI", "EPSON", "HP", "DELL", "ZEBRA", "RICOH", "LENOVO", "SUPER MICRO", "SUPERMICRO"]
    for marca in marcas:
        if marca in texto:
            return "Supermicro" if marca in ("SUPER MICRO", "SUPERMICRO") else marca.title()
    return None

def gerar_codigo_rastreio(cur, tipo_equipamento):
    prefixo = PREFIXOS_RASTREIO.get(tipo_equipamento, "EQP")
    cur.execute(
        "SELECT codigo FROM equipamentos WHERE tipo_equipamento=%s AND codigo LIKE %s",
        (tipo_equipamento, f"RCH-{prefixo}-%")
    )
    numeros = []
    for row in cur.fetchall():
        try:
            numeros.append(int(row['codigo'].split("-")[-1]))
        except (TypeError, ValueError, IndexError, AttributeError):
            pass
    proximo_numero = (max(numeros) if numeros else 0) + 1
    codigo = f"RCH-{prefixo}-{proximo_numero:06d}"

    while True:
        cur.execute("SELECT id FROM equipamentos WHERE codigo=%s", (codigo,))
        if not cur.fetchone():
            return codigo
        proximo_numero += 1
        codigo = f"RCH-{prefixo}-{proximo_numero:06d}"


def _chave_estoque(tipo_suprimento, modelo_impressora):
    tipo = (tipo_suprimento or '').strip()
    if tipo == 'Papel Fotografico':
        return tipo, '-'
    modelo = (modelo_impressora or '').strip().upper()
    return tipo, modelo


def buscar_ou_criar_estoque(cur, tipo_suprimento, modelo_impressora):
    tipo, modelo = _chave_estoque(tipo_suprimento, modelo_impressora)
    cur.execute(
        "SELECT id, quantidade FROM estoque WHERE tipo_suprimento=%s AND modelo_impressora=%s",
        (tipo, modelo)
    )
    row = cur.fetchone()
    if row:
        return row['id'], row['quantidade']
    cur.execute(
        "INSERT INTO estoque (tipo_suprimento, modelo_impressora, quantidade) VALUES (%s, %s, 0) RETURNING id",
        (tipo, modelo)
    )
    return cur.fetchone()['id'], 0


def movimentar_estoque(cur, estoque_id, tipo_movimento, quantidade, saldo_antes, motivo=None, responsavel=None, entrega_id=None):
    saldo_depois = saldo_antes + quantidade
    cur.execute("""
        INSERT INTO estoque_movimentacoes
        (estoque_id, tipo_movimento, quantidade, saldo_antes, saldo_depois, motivo, responsavel, entrega_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (estoque_id, tipo_movimento, quantidade, saldo_antes, saldo_depois, motivo, responsavel, entrega_id))
    cur.execute(
        "UPDATE estoque SET quantidade=%s, data_atualizacao=CURRENT_TIMESTAMP WHERE id=%s",
        (saldo_depois, estoque_id)
    )


def verificar_saldo(cur, tipo_suprimento, modelo_impressora, quantidade):
    tipo, modelo = _chave_estoque(tipo_suprimento, modelo_impressora)
    cur.execute(
        "SELECT quantidade FROM estoque WHERE tipo_suprimento=%s AND modelo_impressora=%s",
        (tipo, modelo)
    )
    row = cur.fetchone()
    saldo = row['quantidade'] if row else 0
    return saldo >= quantidade, saldo


def debitar_estoque_entrega(cur, entrega_id, itens, responsavel=None):
    """Recebe lista de dicts: {tipo_suprimento, modelo_impressora, quantidade}."""
    faltantes = []
    for item in itens:
        tipo, modelo = _chave_estoque(item.get('tipo_suprimento'), item.get('modelo_impressora'))
        qtd = int(item.get('quantidade') or 1)
        ok, saldo = verificar_saldo(cur, tipo, modelo, qtd)
        if not ok:
            faltantes.append({'tipo': tipo, 'modelo': modelo, 'saldo': saldo, 'solicitado': qtd})
    if faltantes:
        return faltantes

    for item in itens:
        tipo, modelo = _chave_estoque(item.get('tipo_suprimento'), item.get('modelo_impressora'))
        qtd = int(item.get('quantidade') or 1)
        estoque_id, saldo = buscar_ou_criar_estoque(cur, tipo, modelo)
        movimentar_estoque(
            cur, estoque_id, 'saida', -qtd, saldo,
            motivo=f'Entrega para unidade (entrega_id={entrega_id})',
            responsavel=responsavel,
            entrega_id=entrega_id
        )
    return []


def estornar_estoque_entrega(cur, entrega_id, responsavel=None):
    cur.execute(
        "SELECT tipo_suprimento, modelo_impressora, quantidade FROM suprimentos_itens WHERE entrega_id=%s",
        (entrega_id,)
    )
    itens = cur.fetchall()
    for item in itens:
        tipo, modelo = _chave_estoque(item['tipo_suprimento'], item['modelo_impressora'])
        qtd = int(item['quantidade'] or 1)
        estoque_id, saldo = buscar_ou_criar_estoque(cur, tipo, modelo)
        movimentar_estoque(
            cur, estoque_id, 'entrada', qtd, saldo,
            motivo=f'Estorno por exclusao de entrega (entrega_id={entrega_id})',
            responsavel=responsavel,
            entrega_id=entrega_id
        )
        cur.execute("UPDATE suprimentos_itens SET estorno_estoque=1 WHERE entrega_id=%s", (entrega_id,))


def importar_planilha_para_banco(caminho=PLANILHA_PADRAO, limpar=True):
    import openpyxl

    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    db = get_db()
    cur = db.cursor()

    if limpar:
        cur.execute("DELETE FROM movimentacoes")
        cur.execute("DELETE FROM historico_defeitos")
        cur.execute("DELETE FROM equipamentos")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    importados = {"impressoras": 0, "computadores": 0, "movimentacoes": 0, "defeitos": 0}

    if "IMPRESSORA" in wb.sheetnames:
        ws = wb["IMPRESSORA"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            data, modelo, serie, contador, movimento, local_coleta, local_atual, motivo, item_id = row[:9]
            modelo = _normalizar_texto(modelo)
            serie = _normalizar_texto(serie)
            if not modelo and not serie:
                continue
            if not modelo:
                modelo = "Modelo não informado"
            local_atual = _normalizar_texto(local_atual) or "RENCH - Estoque"
            local_coleta = _normalizar_texto(local_coleta)
            contador = int(contador) if isinstance(contador, (int, float)) else 0

            cur.execute("SELECT id FROM equipamentos WHERE numero_serie=%s", (serie,))
            existente = cur.fetchone() if serie else None
            if existente:
                equipamento_id = existente['id']
                cur.execute("""
                    UPDATE equipamentos
                    SET modelo=%s, fabricante=COALESCE(fabricante, %s), local_atual_nome=%s,
                        contador_mono=%s, tipo_equipamento='impressora'
                    WHERE id=%s
                """, (modelo, _descobrir_fabricante(modelo), local_atual, contador, equipamento_id))
            else:
                cur.execute("""
                    INSERT INTO equipamentos (
                        codigo, tipo_equipamento, fabricante, modelo, numero_serie,
                        funcao, tipo_impressao, contador_mono, local_atual_nome,
                        status, observacoes
                    ) VALUES (%s, 'impressora', %s, %s, %s, 'impressora', 'laser', %s, %s, 'ativo', %s)
                    RETURNING id
                """, (gerar_codigo_rastreio(cur, "impressora"), _descobrir_fabricante(modelo), modelo, serie, contador, local_atual, _normalizar_texto(motivo)))
                equipamento_id = cur.fetchone()['id']
                importados["impressoras"] += 1

            if data or movimento or local_coleta or local_atual:
                cur.execute("""
                    INSERT INTO movimentacoes (
                        equipamento_id, data_movimentacao, origem_local, destino_local,
                        tipo_movimento, observacoes
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    equipamento_id,
                    _normalizar_data(data) or datetime.now().strftime("%Y-%m-%d"),
                    local_coleta,
                    local_atual,
                    _mapear_tipo_movimento(movimento),
                    _normalizar_texto(movimento) or _normalizar_texto(motivo)
                ))
                importados["movimentacoes"] += 1

    if "COMP-NOTE" in wb.sheetnames:
        ws = wb["COMP-NOTE"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            data, tipo, modelo, serie, movimento, local_coleta, local_atual, motivo, item_id, config = row[:10]
            modelo = _normalizar_texto(modelo)
            serie = _normalizar_texto(serie)
            if not modelo and not serie:
                continue
            if not modelo:
                modelo = "Modelo não informado"
            tipo_equipamento = _mapear_tipo_computador(tipo)
            local_atual = _normalizar_texto(local_atual) or "RENCH - Estoque"
            local_coleta = _normalizar_texto(local_coleta)

            cur.execute("SELECT id FROM equipamentos WHERE numero_serie=%s", (serie,))
            existente = cur.fetchone() if serie else None
            if existente:
                equipamento_id = existente['id']
                cur.execute("""
                    UPDATE equipamentos
                    SET modelo=%s, fabricante=COALESCE(fabricante, %s), local_atual_nome=%s,
                        tipo_equipamento=%s
                    WHERE id=%s
                """, (modelo, _descobrir_fabricante(modelo), local_atual, tipo_equipamento, equipamento_id))
            else:
                cur.execute("""
                    INSERT INTO equipamentos (
                        codigo, tipo_equipamento, fabricante, modelo, numero_serie,
                        local_atual_nome, status, observacoes
                    ) VALUES (%s, %s, %s, %s, %s, %s, 'ativo', %s)
                    RETURNING id
                """, (gerar_codigo_rastreio(cur, tipo_equipamento), tipo_equipamento, _descobrir_fabricante(modelo), modelo, serie, local_atual, _normalizar_texto(config) or _normalizar_texto(motivo)))
                equipamento_id = cur.fetchone()['id']
                importados["computadores"] += 1

            if data or movimento or local_coleta or local_atual:
                cur.execute("""
                    INSERT INTO movimentacoes (
                        equipamento_id, data_movimentacao, origem_local, destino_local,
                        tipo_movimento, observacoes
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    equipamento_id,
                    _normalizar_data(data) or datetime.now().strftime("%Y-%m-%d"),
                    local_coleta,
                    local_atual,
                    _mapear_tipo_movimento(movimento),
                    _normalizar_texto(movimento) or _normalizar_texto(motivo)
                ))
                importados["movimentacoes"] += 1

    if "Planilha1" in wb.sheetnames:
        ws = wb["Planilha1"]
        for row in ws.iter_rows(values_only=True):
            serie = _normalizar_texto(row[0] if len(row) > 0 else None)
            defeito = _normalizar_texto(row[1] if len(row) > 1 else None)
            if not serie or not defeito:
                continue
            cur.execute("SELECT id FROM equipamentos WHERE numero_serie=%s", (serie,))
            equip = cur.fetchone()
            if equip:
                cur.execute("""
                    INSERT INTO historico_defeitos (equipamento_id, descricao)
                    VALUES (%s, %s)
                """, (equip['id'], defeito))
                importados["defeitos"] += 1

    db.commit()
    cur.close()
    return importados

# ============================================================
# ROTAS
# ============================================================

USUARIO_PADRAO = 'admin'
SENHA_PADRAO_HASH = hashlib.sha256('ipascnma'.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logado'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '')
        senha_hash = hashlib.sha256(senha.encode()).hexdigest()
        if usuario == USUARIO_PADRAO and senha_hash == SENHA_PADRAO_HASH:
            session['logado'] = True
            session['usuario'] = usuario
            return redirect(url_for('index'))
        flash('Usuário ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT tipo_equipamento, COUNT(*) as qtd FROM equipamentos WHERE ativo=1 GROUP BY tipo_equipamento")
    por_tipo = {r['tipo_equipamento']: r['qtd'] for r in cur.fetchall()}

    cur.execute("SELECT local_atual_nome, COUNT(*) as qtd FROM equipamentos WHERE ativo=1 AND local_atual_nome IS NOT NULL GROUP BY local_atual_nome")
    por_local = cur.fetchall()

    cur.execute("""
        SELECT m.data_movimentacao, e.tipo_equipamento, e.modelo, e.numero_serie, m.tipo_movimento, m.destino_local, m.responsavel
        FROM movimentacoes m
        JOIN equipamentos e ON m.equipamento_id = e.id
        WHERE m.data_movimentacao LIKE '____-__-__'
        ORDER BY m.data_movimentacao DESC, m.id DESC
        LIMIT 8
    """)
    ultimas_mov = cur.fetchall()

    return render_template('dashboard.html',
        por_tipo=por_tipo, por_local=por_local, ultimas_mov=ultimas_mov
    )

@app.route('/importar-planilha', methods=['POST'])
@login_required
def importar_planilha():
    try:
        resultado = importar_planilha_para_banco()
        flash(
            f"Planilha importada: {resultado['impressoras']} impressoras, "
            f"{resultado['computadores']} computadores/notebooks/servidores, "
            f"{resultado['movimentacoes']} movimentações e {resultado['defeitos']} defeitos.",
            "success"
        )
    except Exception as erro:
        flash(f"Erro ao importar planilha: {erro}", "danger")
    return redirect(url_for('lista_equipamentos'))

@app.route('/equipamentos')
@login_required
def lista_equipamentos():
    db = get_db()
    cur = db.cursor()
    tipo = request.args.get('tipo', '')
    busca = request.args.get('q', '')
    mov_recente = request.args.get('mov_recente', '')
    unidade_id = request.args.get('unidade_id', '')
    empresa_id = request.args.get('empresa_id', '')
    sugestoes = []

    sql = """
        SELECT e.*, COALESCE(u.nome, e.local_atual_nome) as local_nome,
               emp.nome as empresa_nome, emp.id as empresa_id, u.setor as unidade_setor,
               m.ultima_movimentacao
        FROM equipamentos e
        LEFT JOIN unidades u ON e.unidade_id = u.id
        LEFT JOIN empresas emp ON emp.id = u.empresa_id
        LEFT JOIN (
            SELECT equipamento_id, MAX(data_movimentacao) as ultima_movimentacao
            FROM movimentacoes
            WHERE data_movimentacao LIKE '____-__-__'
            GROUP BY equipamento_id
        ) m ON m.equipamento_id = e.id
        WHERE e.ativo=1
    """
    params = []

    if tipo:
        sql += " AND e.tipo_equipamento = %s"
        params.append(tipo)
    if unidade_id:
        sql += " AND e.unidade_id = %s"
        params.append(unidade_id)
    if empresa_id == 'sem_empresa':
        sql += " AND emp.id IS NULL"
    elif empresa_id:
        sql += " AND emp.id = %s"
        params.append(empresa_id)
    sql += " ORDER BY e.tipo_equipamento, e.modelo"

    cur.execute(sql, params)
    equipamentos = cur.fetchall()

    if busca:
        termos = preparar_termos_busca(busca)
        filtrados = []
        for equipamento in equipamentos:
            pontuacao = calcular_pontuacao_busca(
                termos,
                equipamento["codigo"],
                equipamento["fabricante"],
                equipamento["modelo"],
                equipamento["numero_serie"],
                equipamento["patrimonio"],
                equipamento["cliente_atual"],
                equipamento["local_nome"],
                equipamento["empresa_nome"],
                equipamento["unidade_setor"],
                equipamento.get("setor_equipamento"),
            )
            if pontuacao >= 50:
                filtrados.append((pontuacao, equipamento))

        filtrados.sort(key=lambda item: item[0], reverse=True)
        equipamentos = [item[1] for item in filtrados]
        sugestoes = buscar_sugestoes_locais(cur, busca)

    cur.execute("""
        SELECT u.id, u.nome, e.nome as empresa_nome
        FROM unidades u
        JOIN empresas e ON e.id = u.empresa_id
        WHERE u.ativo=1
        ORDER BY e.nome, u.nome
    """)
    unidades_filtro = cur.fetchall()

    cur.execute("""
        SELECT emp.id, emp.nome, COUNT(eq.id) as qtd
        FROM empresas emp
        LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
        LEFT JOIN equipamentos eq ON eq.unidade_id = u.id AND eq.ativo=1
        WHERE emp.ativo=1
        GROUP BY emp.id, emp.nome
        ORDER BY emp.nome
    """)
    empresas_filtro = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) as qtd FROM equipamentos e
        LEFT JOIN unidades u ON e.unidade_id = u.id
        WHERE e.ativo=1 AND (e.unidade_id IS NULL OR u.empresa_id IS NULL)
    """)
    qtd_sem_empresa = cur.fetchone()['qtd']

    return render_template('equipamentos.html',
        equipamentos=equipamentos, filtro_tipo=tipo, filtro_mov=mov_recente, busca=busca,
        filtro_unidade_id=unidade_id, unidades_filtro=unidades_filtro, sugestoes=sugestoes,
        filtro_empresa_id=empresa_id, empresas_filtro=empresas_filtro, qtd_sem_empresa=qtd_sem_empresa
    )

@app.route('/equipamento/<int:equip_id>')
@login_required
def detalhe_equipamento(equip_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM equipamentos WHERE id=%s AND ativo=1", (equip_id,))
    equip = cur.fetchone()
    if not equip:
        flash("Equipamento nao encontrado!", "danger")
        return redirect(url_for('lista_equipamentos'))

    cur.execute("SELECT * FROM movimentacoes WHERE equipamento_id=%s ORDER BY data_movimentacao DESC, id DESC", (equip_id,))
    movimentacoes = cur.fetchall()

    cur.execute("SELECT * FROM historico_defeitos WHERE equipamento_id=%s ORDER BY data_ocorrencia DESC", (equip_id,))
    defeitos = cur.fetchall()

    return render_template('equipamento_detalhe.html',
        equip=equip, movimentacoes=movimentacoes, defeitos=defeitos
    )

@app.route('/equipamento/novo', methods=['GET', 'POST'])
@login_required
def novo_equipamento():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        tipo = request.form['tipo_equipamento']
        unidade_id = request.form.get('unidade_id') or None
        local_atual_nome = request.form.get('local_atual_nome')
        if unidade_id:
            cur.execute("SELECT u.nome, e.nome FROM unidades u JOIN empresas e ON e.id=u.empresa_id WHERE u.id=%s", (unidade_id,))
            unidade = cur.fetchone()
            if unidade:
                local_atual_nome = unidade[0]
                cliente_atual = unidade[1]
            else:
                cliente_atual = None
        else:
            cliente_atual = None

        campos = {
            'codigo': gerar_codigo_rastreio(cur, tipo),
            'tipo_equipamento': tipo,
            'fabricante': request.form.get('fabricante'),
            'modelo': request.form.get('modelo'),
            'numero_serie': request.form.get('numero_serie'),
            'patrimonio': request.form.get('patrimonio'),
            'setor_equipamento': request.form.get('setor_equipamento'),
            'status': request.form.get('status', 'ativo'),
            'unidade_id': unidade_id,
            'local_atual_nome': local_atual_nome,
            'cliente_atual': cliente_atual,
            'observacoes': request.form.get('observacoes'),
        }

        if tipo == 'impressora':
            campos.update({
                'funcao': request.form.get('funcao'),
                'tipo_impressao': request.form.get('tipo_impressao'),
                'tamanho_papel': request.form.get('tamanho_papel'),
                'funcionalidades': request.form.get('funcionalidades'),
                'contador_mono': request.form.get('contador_mono', 0),
                'contador_color': request.form.get('contador_color', 0),
            })
        else:
            campos.update({
                'processador_modelo': request.form.get('processador_modelo'),
                'processador_geracao': request.form.get('processador_geracao'),
                'processador_velocidade': request.form.get('processador_velocidade'),
                'memoria_capacidade': request.form.get('memoria_capacidade'),
                'memoria_tipo': request.form.get('memoria_tipo'),
                'memoria_velocidade': request.form.get('memoria_velocidade'),
                'armazenamento_1_capacidade': request.form.get('armazenamento_1_capacidade'),
                'armazenamento_1_tipo': request.form.get('armazenamento_1_tipo'),
                'armazenamento_2_capacidade': request.form.get('armazenamento_2_capacidade'),
                'armazenamento_2_tipo': request.form.get('armazenamento_2_tipo'),
                'armazenamento_3_capacidade': request.form.get('armazenamento_3_capacidade'),
                'armazenamento_3_tipo': request.form.get('armazenamento_3_tipo'),
            })

        colunas = [k for k, v in campos.items() if v is not None]
        valores = [campos[k] for k in colunas]
        placeholders = ','.join(['%s' for _ in colunas])

        sql = f"INSERT INTO equipamentos ({','.join(colunas)}) VALUES ({placeholders})"
        cur.execute(sql, valores)
        db.commit()

        flash("Equipamento cadastrado com sucesso!", "success")
        return redirect(url_for('lista_equipamentos'))

    cur.execute("""
        SELECT e.id as empresa_id, e.nome as empresa_nome, e.tipo as empresa_tipo,
               u.id as unidade_id, u.nome as unidade_nome, u.setor
        FROM empresas e
        LEFT JOIN unidades u ON u.empresa_id = e.id AND u.ativo=1
        WHERE e.ativo=1
        ORDER BY e.tipo DESC, e.nome, u.nome
    """)
    locais = cur.fetchall()

    return render_template('equipamento_form.html', equip=None, locais=locais)

@app.route('/equipamento/<int:equip_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_equipamento(equip_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM equipamentos WHERE id=%s AND ativo=1", (equip_id,))
    equip = cur.fetchone()
    if not equip:
        flash("Equipamento nao encontrado!", "danger")
        return redirect(url_for('lista_equipamentos'))

    if request.method == 'POST':
        unidade_id = request.form.get('unidade_id') or None
        local_atual_nome = request.form.get('local_atual_nome')
        cliente_atual = None

        if unidade_id:
            cur.execute("""
                SELECT u.nome as unidade_nome, e.nome as empresa_nome
                FROM unidades u
                JOIN empresas e ON e.id = u.empresa_id
                WHERE u.id=%s
            """, (unidade_id,))
            unidade = cur.fetchone()
            if unidade:
                local_atual_nome = unidade['unidade_nome']
                cliente_atual = unidade['empresa_nome']

        campos = {
            'fabricante': request.form.get('fabricante'),
            'modelo': request.form.get('modelo'),
            'numero_serie': request.form.get('numero_serie'),
            'patrimonio': request.form.get('patrimonio'),
            'setor_equipamento': request.form.get('setor_equipamento'),
            'status': request.form.get('status', 'ativo'),
            'unidade_id': unidade_id,
            'local_atual_nome': local_atual_nome,
            'cliente_atual': cliente_atual,
            'observacoes': request.form.get('observacoes'),
            'funcao': request.form.get('funcao') or None,
            'tipo_impressao': request.form.get('tipo_impressao') or None,
            'tamanho_papel': request.form.get('tamanho_papel') or None,
            'funcionalidades': request.form.get('funcionalidades'),
            'contador_mono': request.form.get('contador_mono') or 0,
            'contador_color': request.form.get('contador_color') or 0,
            'processador_modelo': request.form.get('processador_modelo'),
            'processador_geracao': request.form.get('processador_geracao'),
            'processador_velocidade': request.form.get('processador_velocidade'),
            'memoria_capacidade': request.form.get('memoria_capacidade'),
            'memoria_tipo': request.form.get('memoria_tipo'),
            'memoria_velocidade': request.form.get('memoria_velocidade'),
            'armazenamento_1_capacidade': request.form.get('armazenamento_1_capacidade'),
            'armazenamento_1_tipo': request.form.get('armazenamento_1_tipo'),
            'armazenamento_2_capacidade': request.form.get('armazenamento_2_capacidade'),
            'armazenamento_2_tipo': request.form.get('armazenamento_2_tipo'),
            'armazenamento_3_capacidade': request.form.get('armazenamento_3_capacidade'),
            'armazenamento_3_tipo': request.form.get('armazenamento_3_tipo'),
        }

        set_sql = ', '.join([f"{campo}=%s" for campo in campos])
        valores = list(campos.values()) + [equip_id]
        cur.execute(f"UPDATE equipamentos SET {set_sql} WHERE id=%s", valores)
        db.commit()

        flash("Equipamento atualizado com sucesso!", "success")
        return redirect(url_for('detalhe_equipamento', equip_id=equip_id))

    cur.execute("""
        SELECT e.id as empresa_id, e.nome as empresa_nome, e.tipo as empresa_tipo,
               u.id as unidade_id, u.nome as unidade_nome, u.setor
        FROM empresas e
        LEFT JOIN unidades u ON u.empresa_id = e.id AND u.ativo=1
        WHERE e.ativo=1
        ORDER BY e.tipo DESC, e.nome, u.nome
    """)
    locais = cur.fetchall()

    return render_template('equipamento_editar.html', equip=equip, locais=locais)

@app.route('/equipamentos-por-unidade')
@login_required
def equipamentos_por_unidade():
    db = get_db()
    cur = db.cursor()

    empresa_id = request.args.get('empresa_id', '')
    unidade_id = request.args.get('unidade_id', '')
    tipo = request.args.get('tipo', '')

    cur.execute("SELECT id, nome FROM empresas WHERE ativo=1 ORDER BY nome")
    empresas = cur.fetchall()

    unidades = []
    equipamentos = []
    unidade_selecionada = None

    if empresa_id:
        cur.execute("""
            SELECT id, nome, setor FROM unidades
            WHERE empresa_id=%s AND ativo=1
            ORDER BY nome
        """, (empresa_id,))
        unidades = cur.fetchall()

    if unidade_id:
        cur.execute("SELECT u.*, e.nome as empresa_nome FROM unidades u LEFT JOIN empresas e ON e.id=u.empresa_id WHERE u.id=%s", (unidade_id,))
        unidade_selecionada = cur.fetchone()

        sql = """
            SELECT e.*, COALESCE(u.nome, e.local_atual_nome) as local_nome,
                   emp.nome as empresa_nome
            FROM equipamentos e
            LEFT JOIN unidades u ON e.unidade_id = u.id
            LEFT JOIN empresas emp ON emp.id = u.empresa_id
            WHERE e.ativo=1 AND (e.unidade_id = %s OR e.local_atual_nome = (SELECT nome FROM unidades WHERE id=%s))
        """
        params = [unidade_id, unidade_id]
        if tipo:
            sql += " AND e.tipo_equipamento = %s"
            params.append(tipo)
        sql += " ORDER BY e.tipo_equipamento, e.modelo"
        cur.execute(sql, params)
        equipamentos = cur.fetchall()

    return render_template('equipamentos_por_unidade.html',
        empresas=empresas, unidades=unidades, equipamentos=equipamentos,
        filtro_empresa_id=empresa_id, filtro_unidade_id=unidade_id,
        filtro_tipo=tipo, unidade_selecionada=unidade_selecionada
    )


@app.route('/equipamento/<int:equip_id>/excluir', methods=['POST'])
@login_required
def excluir_equipamento(equip_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("UPDATE equipamentos SET ativo=0 WHERE id=%s", (equip_id,))
    db.commit()
    flash('Equipamento excluido com sucesso.', 'success')
    return redirect(url_for('lista_equipamentos'))

@app.route('/movimentar/<int:equip_id>', methods=['GET', 'POST'])
@login_required
def movimentar(equip_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT e.*, u.nome as unidade_nome, emp.nome as empresa_nome FROM equipamentos e LEFT JOIN unidades u ON u.id=e.unidade_id LEFT JOIN empresas emp ON emp.id=u.empresa_id WHERE e.id=%s", (equip_id,))
    equip = cur.fetchone()
    if not equip:
        flash("Equipamento nao encontrado!", "danger")
        return redirect(url_for('lista_equipamentos'))

    if request.method == 'POST':
        tipo_mov = request.form['tipo_movimento']
        data_mov = request.form['data_movimentacao']
        destino_unidade_id = request.form.get('destino_unidade_id')
        responsavel = request.form.get('responsavel')
        obs = request.form.get('observacoes')
        setor_destino = request.form.get('setor_equipamento', '').strip() or None
        contador_mono_novo = request.form.get('contador_mono_novo', '').strip()
        contador_color_novo = request.form.get('contador_color_novo', '').strip()

        contador_mono_anterior = int(equip['contador_mono'] or 0)
        contador_color_anterior = int(equip['contador_color'] or 0)
        contador_mono_novo_int = int(contador_mono_novo) if contador_mono_novo else contador_mono_anterior
        contador_color_novo_int = int(contador_color_novo) if contador_color_novo else contador_color_anterior

        destino_unidade_nome = None
        if destino_unidade_id:
            cur.execute("SELECT nome, empresa_id FROM unidades WHERE id=%s", (destino_unidade_id,))
            unidade_dest = cur.fetchone()
            if unidade_dest:
                destino_unidade_nome = unidade_dest['nome']

        cur.execute("""
            INSERT INTO movimentacoes (equipamento_id, data_movimentacao, tipo_movimento,
                origem_local, origem_unidade, destino_local, destino_unidade, responsavel, observacoes,
                contador_mono_anterior, contador_mono_novo, contador_color_anterior, contador_color_novo)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (equip_id, data_mov, tipo_mov,
              equip['local_atual_nome'], equip['unidade_nome'] or equip['local_atual_nome'],
              destino_unidade_nome or "Sem unidade", destino_unidade_nome,
              responsavel, obs,
              contador_mono_anterior, contador_mono_novo_int,
              contador_color_anterior, contador_color_novo_int))

        cur.execute("""
            UPDATE equipamentos SET unidade_id=%s, local_atual_nome=%s, cliente_atual=%s,
                contador_mono=%s, contador_color=%s, setor_equipamento=%s WHERE id=%s
        """, (destino_unidade_id, destino_unidade_nome, None,
              contador_mono_novo_int, contador_color_novo_int, setor_destino, equip_id))

        db.commit()
        flash("Movimentacao registrada com sucesso!", "success")
        return redirect(url_for('detalhe_equipamento', equip_id=equip_id))

    cur.execute("""
        SELECT e.id as empresa_id, e.nome as empresa_nome, e.tipo as empresa_tipo,
               u.id as unidade_id, u.nome as unidade_nome, u.setor
        FROM empresas e
        LEFT JOIN unidades u ON u.empresa_id = e.id AND u.ativo=1
        WHERE e.ativo=1
        ORDER BY e.tipo DESC, e.nome, u.nome
    """)
    locais = cur.fetchall()

    cur.execute("""
        SELECT data_movimentacao, tipo_movimento, destino_local, destino_unidade,
               contador_mono_anterior, contador_mono_novo, contador_color_anterior, contador_color_novo
        FROM movimentacoes
        WHERE equipamento_id=%s
        ORDER BY data_movimentacao DESC, id DESC
        LIMIT 20
    """, (equip_id,))
    historico_contadores = cur.fetchall()

    return render_template('movimentacao_form.html', equip=equip, locais=locais, historico_contadores=historico_contadores)

@app.route('/locais', methods=['GET', 'POST'])
@login_required
def lista_locais():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        if request.form.get('acao') == 'empresa':
            nome = request.form.get('nome')
            tipo = request.form.get('tipo', 'cliente')
            if nome:
                cur.execute("INSERT INTO empresas (nome, tipo) VALUES (%s, %s)", (nome, tipo))
                db.commit()
                flash("Empresa cadastrada com sucesso!", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'empresa_editar':
            empresa_id = request.form.get('empresa_id')
            nome = request.form.get('nome')
            tipo = request.form.get('tipo', 'cliente')
            if empresa_id and nome:
                cur.execute("UPDATE empresas SET nome=%s, tipo=%s WHERE id=%s", (nome, tipo, empresa_id))
                db.commit()
                flash("Empresa atualizada com sucesso!", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'empresa_excluir':
            empresa_id = request.form.get('empresa_id')
            if empresa_id:
                cur.execute("UPDATE empresas SET ativo=0 WHERE id=%s", (empresa_id,))
                cur.execute("UPDATE unidades SET ativo=0 WHERE empresa_id=%s", (empresa_id,))
                db.commit()
                flash("Empresa excluida com sucesso.", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'unidade':
            empresa_id = request.form.get('empresa_id')
            nome = request.form.get('unidade_nome')
            setor = request.form.get('setor')
            if empresa_id and nome:
                cur.execute("INSERT INTO unidades (empresa_id, nome, setor) VALUES (%s, %s, %s)", (empresa_id, nome, setor))
                db.commit()
                flash("Unidade cadastrada com sucesso!", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'unidade_editar':
            unidade_id = request.form.get('unidade_id')
            nome = request.form.get('unidade_nome')
            setor = request.form.get('setor')
            empresa_id = request.form.get('empresa_id')
            if unidade_id and nome and empresa_id:
                cur.execute("UPDATE unidades SET nome=%s, setor=%s, empresa_id=%s WHERE id=%s", (nome, setor, empresa_id, unidade_id))
                db.commit()
                flash("Unidade atualizada com sucesso!", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'unidade_excluir':
            unidade_id = request.form.get('unidade_id')
            if unidade_id:
                cur.execute("UPDATE unidades SET ativo=0 WHERE id=%s", (unidade_id,))
                db.commit()
                flash("Unidade excluida com sucesso.", "success")
            return redirect(url_for('lista_locais'))

        if request.form.get('acao') == 'unidade_unificar':
            manter_id = request.form.get('manter_id')
            remover_id = request.form.get('remover_id')
            if manter_id and remover_id and manter_id != remover_id:
                cur.execute("UPDATE equipamentos SET unidade_id=%s WHERE unidade_id=%s", (manter_id, remover_id))
                cur.execute("""
                    UPDATE equipamentos SET local_atual_nome=(SELECT nome FROM unidades WHERE id=%s)
                    WHERE unidade_id=%s AND local_atual_nome=(SELECT nome FROM unidades WHERE id=%s)
                """, (manter_id, manter_id, remover_id))
                cur.execute("UPDATE unidades SET ativo=0 WHERE id=%s", (remover_id,))
                db.commit()
                flash("Unidades unificadas com sucesso!", "success")
            return redirect(url_for('lista_locais'))

    cur.execute("""
        SELECT e.id as empresa_id, e.nome as empresa_nome, e.tipo as empresa_tipo,
               u.id as unidade_id, u.nome as unidade_nome, u.setor,
               COALESCE(eq.qtd, 0) as qtd_equipamentos
        FROM empresas e
        LEFT JOIN unidades u ON u.empresa_id = e.id AND u.ativo=1
        LEFT JOIN (
            SELECT unidade_id, COUNT(*) as qtd FROM equipamentos WHERE ativo=1 AND unidade_id IS NOT NULL GROUP BY unidade_id
        ) eq ON eq.unidade_id = u.id
        WHERE e.ativo=1
        ORDER BY e.tipo DESC, e.nome, u.nome
    """)
    locais = cur.fetchall()
    cur.execute("SELECT id, nome, tipo FROM empresas WHERE ativo=1 ORDER BY tipo DESC, nome")
    empresas = cur.fetchall()

    cur.execute("""
        SELECT u.id, u.nome, u.setor, e.nome as empresa_nome
        FROM unidades u
        JOIN empresas e ON e.id = u.empresa_id
        WHERE u.ativo=1 AND e.ativo=1
        ORDER BY e.nome, u.nome
    """)
    todas_unidades = cur.fetchall()

    return render_template('locais.html', locais=locais, empresas=empresas, todas_unidades=todas_unidades)

@app.route('/acesso-remoto', methods=['GET', 'POST'])
@login_required
def acesso_remoto():
    db = get_db()
    cur = db.cursor()
    busca = request.args.get('q', '')

    if request.method == 'POST':
        acao = request.form.get('acao')
        if acao == 'novo':
            cur.execute("""
                INSERT INTO acesso_remoto (local, anydesk, senha, observacoes)
                VALUES (%s, %s, %s, %s)
            """, (
                request.form.get('local'),
                request.form.get('anydesk'),
                request.form.get('senha'),
                request.form.get('observacoes')
            ))
            db.commit()
            flash('Acesso remoto cadastrado com sucesso!', 'success')
        elif acao == 'editar':
            cur.execute("""
                UPDATE acesso_remoto
                SET local=%s, anydesk=%s, senha=%s, observacoes=%s
                WHERE id=%s
            """, (
                request.form.get('local'),
                request.form.get('anydesk'),
                request.form.get('senha'),
                request.form.get('observacoes'),
                request.form.get('id')
            ))
            db.commit()
            flash('Acesso remoto atualizado com sucesso!', 'success')
        elif acao == 'excluir':
            cur.execute("UPDATE acesso_remoto SET ativo=0 WHERE id=%s", (request.form.get('id'),))
            db.commit()
            flash('Acesso remoto excluido.', 'success')
        return redirect(url_for('acesso_remoto', q=busca))

    sql = "SELECT * FROM acesso_remoto WHERE ativo=1"
    params = []
    if busca:
        sql += " AND (local ILIKE %s OR anydesk ILIKE %s OR senha ILIKE %s OR observacoes ILIKE %s)"
        like = f'%{busca}%'
        params = [like, like, like, like]
    sql += " ORDER BY local"
    cur.execute(sql, params)
    acessos = cur.fetchall()
    return render_template('acesso_remoto.html', acessos=acessos, busca=busca)

@app.route('/historico')
@login_required
def historico():
    db = get_db()
    cur = db.cursor()

    equip_id = request.args.get('equipamento_id')
    serie = request.args.get('serie')

    equip = None
    movimentacoes = []
    sugestoes = []

    if equip_id:
        cur.execute("SELECT * FROM equipamentos WHERE id=%s", (equip_id,))
        equip = cur.fetchone()
        if equip:
            cur.execute("""
                SELECT m.*, e.modelo, e.numero_serie
                FROM movimentacoes m
                JOIN equipamentos e ON m.equipamento_id = e.id
                WHERE m.equipamento_id = %s
                ORDER BY m.data_movimentacao DESC
            """, (equip_id,))
            movimentacoes = cur.fetchall()
    elif serie:
        cur.execute("SELECT e.*, u.nome as unidade_nome, emp.nome as empresa_nome FROM equipamentos e LEFT JOIN unidades u ON u.id=e.unidade_id LEFT JOIN empresas emp ON emp.id=u.empresa_id WHERE e.ativo=1")
        candidatos = cur.fetchall()
        termos = preparar_termos_busca(serie)
        encontrados = []
        for candidato in candidatos:
            pontuacao = calcular_pontuacao_busca(
                termos,
                candidato["numero_serie"],
                candidato["codigo"],
                candidato["modelo"],
                candidato["fabricante"],
                candidato["local_atual_nome"],
                candidato["cliente_atual"],
                candidato["unidade_nome"],
                candidato["empresa_nome"],
            )
            if pontuacao:
                encontrados.append((pontuacao, candidato))

        encontrados.sort(key=lambda item: item[0], reverse=True)
        if encontrados and encontrados[0][0] >= 50:
            equip = encontrados[0][1]
            sugestoes = [item[1] for item in encontrados[1:6]]
            cur.execute("""
                SELECT m.*, e.modelo, e.numero_serie
                FROM movimentacoes m
                JOIN equipamentos e ON m.equipamento_id = e.id
                WHERE m.equipamento_id = %s
                ORDER BY m.data_movimentacao DESC
            """, (equip['id'],))
            movimentacoes = cur.fetchall()
        else:
            equip = None
            sugestoes = []

    return render_template('historico.html', equip=equip, movimentacoes=movimentacoes, serie=serie, sugestoes=sugestoes)

@app.route('/movimentar-equipamento', methods=['GET', 'POST'])
@login_required
def tela_movimentar():
    db = get_db()
    cur = db.cursor()
    busca = request.args.get('q', '')
    tipo = request.args.get('tipo', '')
    empresa_id = request.args.get('empresa_id', '')
    sugestoes = []
    equipamento = None

    if request.method == 'POST':
        equip_id = request.form.get('equipamento_id')
        if equip_id:
            return redirect(url_for('movimentar', equip_id=equip_id))

    sql = """
        SELECT e.*, COALESCE(u.nome, e.local_atual_nome) as local_nome,
               emp.nome as empresa_nome, emp.id as empresa_id, u.setor as unidade_setor,
               m.ultima_movimentacao
        FROM equipamentos e
        LEFT JOIN unidades u ON e.unidade_id = u.id
        LEFT JOIN empresas emp ON emp.id = u.empresa_id
        LEFT JOIN (
            SELECT equipamento_id, MAX(data_movimentacao) as ultima_movimentacao
            FROM movimentacoes
            WHERE data_movimentacao LIKE '____-__-__'
            GROUP BY equipamento_id
        ) m ON m.equipamento_id = e.id
        WHERE e.ativo=1
    """
    params = []

    if tipo:
        sql += " AND e.tipo_equipamento = %s"
        params.append(tipo)
    if empresa_id == 'sem_empresa':
        sql += " AND emp.id IS NULL"
    elif empresa_id:
        sql += " AND emp.id = %s"
        params.append(empresa_id)
    sql += " ORDER BY e.tipo_equipamento, e.modelo"

    cur.execute(sql, params)
    equipamentos = cur.fetchall()

    if busca:
        termos = preparar_termos_busca(busca)
        filtrados = []
        for eq in equipamentos:
            pontuacao = calcular_pontuacao_busca(
                termos, eq["codigo"], eq["fabricante"], eq["modelo"],
                eq["numero_serie"], eq["patrimonio"], eq["cliente_atual"],
                eq["local_nome"], eq["empresa_nome"], eq["unidade_setor"],
            )
            if pontuacao >= 50:
                filtrados.append((pontuacao, eq))
        filtrados.sort(key=lambda i: i[0], reverse=True)
        equipamentos = [i[1] for i in filtrados]
        sugestoes = buscar_sugestoes_locais(cur, busca)

    cur.execute("""
        SELECT emp.id, emp.nome, COUNT(eq.id) as qtd
        FROM empresas emp
        LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
        LEFT JOIN equipamentos eq ON eq.unidade_id = u.id AND eq.ativo=1
        WHERE emp.ativo=1
        GROUP BY emp.id, emp.nome
        ORDER BY emp.nome
    """)
    empresas_filtro = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) as qtd FROM equipamentos e
        LEFT JOIN unidades u ON e.unidade_id = u.id
        WHERE e.ativo=1 AND (e.unidade_id IS NULL OR u.empresa_id IS NULL)
    """)
    qtd_sem_empresa = cur.fetchone()['qtd']

    return render_template('tela_movimentar.html',
        equipamentos=equipamentos, busca=busca, filtro_tipo=tipo, sugestoes=sugestoes,
        filtro_empresa_id=empresa_id, empresas_filtro=empresas_filtro, qtd_sem_empresa=qtd_sem_empresa)

@app.route('/suprimentos', methods=['GET', 'POST'])
@login_required
def lista_suprimentos():
    db = get_db()
    cur = db.cursor()

    busca = request.args.get('q', '').strip()
    unidade_id = request.args.get('unidade_id', '').strip()
    tipo = request.args.get('tipo', '').strip()

    sql = """
        SELECT se.id, se.data_entrega, se.responsavel, se.observacoes,
               u.nome as unidade_nome, emp.nome as empresa_nome,
               si.id as item_id, si.tipo_suprimento, si.modelo_impressora, si.quantidade,
               si.motivo_padrao, si.defeito, si.motivo
        FROM suprimentos_entregas se
        JOIN unidades u ON u.id = se.unidade_id
        JOIN empresas emp ON emp.id = u.empresa_id
        LEFT JOIN suprimentos_itens si ON si.entrega_id = se.id
        WHERE 1=1
    """
    params = []

    if unidade_id:
        sql += " AND se.unidade_id = %s"
        params.append(unidade_id)
    if tipo:
        sql += " AND si.tipo_suprimento = %s"
        params.append(tipo)
    if busca:
        sql += """ AND (
            unaccent(lower(u.nome)) LIKE unaccent(lower(%s))
            OR unaccent(lower(emp.nome)) LIKE unaccent(lower(%s))
            OR unaccent(lower(si.tipo_suprimento)) LIKE unaccent(lower(%s))
        )"""
        params.extend([f'%{busca}%', f'%{busca}%', f'%{busca}%'])

    sql += " ORDER BY se.data_entrega DESC, se.id DESC, si.id"
    cur.execute(sql, params)
    rows = cur.fetchall()

    entregas = {}
    for r in rows:
        eid = r['id']
        if eid not in entregas:
            entregas[eid] = {
                'id': eid,
                'data_entrega': r['data_entrega'],
                'responsavel': r['responsavel'],
                'observacoes': r['observacoes'],
                'unidade_nome': r['unidade_nome'],
                'empresa_nome': r['empresa_nome'],
                'itens': []
            }
        if r['item_id']:
            entregas[eid]['itens'].append({
                'tipo': r['tipo_suprimento'],
                'modelo': r['modelo_impressora'],
                'quantidade': r['quantidade'],
                'motivo_padrao': r['motivo_padrao'],
                'defeito': r['defeito'],
                'motivo': r['motivo']
            })

    cur.execute("""
        SELECT emp.id as empresa_id, emp.nome as empresa_nome,
               u.id as unidade_id, u.nome as unidade_nome
        FROM empresas emp
        LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
        WHERE emp.ativo=1
        ORDER BY emp.tipo DESC, emp.nome, u.nome
    """)
    locais = cur.fetchall()

    cur.execute("SELECT DISTINCT tipo_suprimento FROM suprimentos_itens ORDER BY tipo_suprimento")
    tipos = [row['tipo_suprimento'] for row in cur.fetchall()]

    return render_template('suprimentos.html',
        entregas=list(entregas.values()), locais=locais, tipos=tipos,
        busca=busca, filtro_unidade=unidade_id, filtro_tipo=tipo)

@app.route('/suprimentos/mobile', methods=['GET', 'POST'])
@login_required
def suprimento_mobile():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        unidade_id = request.form['unidade_id']
        data_entrega = request.form['data_entrega']
        responsavel = request.form.get('responsavel')
        observacoes = request.form.get('observacoes')

        cur.execute("""
            INSERT INTO suprimentos_entregas (unidade_id, data_entrega, responsavel, observacoes)
            VALUES (%s, %s, %s, %s) RETURNING id
        """, (unidade_id, data_entrega, responsavel, observacoes))
        entrega_id = cur.fetchone()['id']

        tipos = request.form.getlist('tipo_suprimento[]')
        modelos = request.form.getlist('modelo_impressora[]')
        quantidades = request.form.getlist('quantidade[]')
        motivos_padrao = request.form.getlist('motivo_padrao[]')
        defeitos = request.form.getlist('defeito[]')
        motivos_outros = request.form.getlist('motivo[]')

        itens = []
        for tipo, modelo, qtd, mp, defeito, outro in zip(tipos, modelos, quantidades, motivos_padrao, defeitos, motivos_outros):
            tipo = (tipo or '').strip()
            if tipo:
                itens.append({
                    'tipo_suprimento': tipo,
                    'modelo_impressora': modelo,
                    'quantidade': int(qtd or 1),
                    'motivo_padrao': mp,
                    'defeito': defeito,
                    'outro': outro
                })

        if itens:
            faltantes = debitar_estoque_entrega(cur, entrega_id, itens, responsavel=responsavel)
            if faltantes:
                db.rollback()
                msgs = []
                for f in faltantes:
                    msgs.append(f"{f['tipo']} {f['modelo']}: solicitado {f['solicitado']}, em estoque {f['saldo']}")
                flash('Estoque insuficiente: ' + '; '.join(msgs), 'danger')
                cur.execute("""
                    SELECT emp.id as empresa_id, emp.nome as empresa_nome,
                           u.id as unidade_id, u.nome as unidade_nome
                    FROM empresas emp
                    LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
                    WHERE emp.ativo=1
                    ORDER BY emp.tipo DESC, emp.nome, u.nome
                """)
                locais = cur.fetchall()
                return render_template('suprimento_mobile.html', locais=locais, hoje=data_entrega, aba='entrega')

        for item in itens:
            mp = item['motivo_padrao']
            defeito = item['defeito']
            outro = item['outro']
            motivo_texto = None
            if mp == 'outro':
                motivo_texto = (outro or '').strip() or None
            elif mp:
                motivo_texto = mp
            cur.execute("""
                INSERT INTO suprimentos_itens (entrega_id, tipo_suprimento, modelo_impressora, quantidade, motivo_padrao, defeito, motivo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (entrega_id, item['tipo_suprimento'], item['modelo_impressora'].strip() or None, item['quantidade'], mp or None, defeito.strip() or None, motivo_texto))

        db.commit()
        flash('Entrega salva com sucesso!', 'success')
        return redirect(url_for('suprimento_mobile'))

    cur.execute("""
        SELECT emp.id as empresa_id, emp.nome as empresa_nome,
               u.id as unidade_id, u.nome as unidade_nome
        FROM empresas emp
        LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
        WHERE emp.ativo=1
        ORDER BY emp.tipo DESC, emp.nome, u.nome
    """)
    locais = cur.fetchall()

    cur.execute("SELECT id, tipo_suprimento, modelo_impressora, quantidade FROM estoque ORDER BY tipo_suprimento, modelo_impressora")
    estoque = cur.fetchall()

    hoje = datetime.now().strftime('%Y-%m-%d')
    return render_template('suprimento_mobile.html', locais=locais, hoje=hoje, estoque=estoque, aba=request.args.get('aba', 'entrega'))

@app.route('/suprimentos/novo', methods=['GET', 'POST'])
@login_required
def novo_suprimento():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        unidade_id = request.form['unidade_id']
        data_entrega = request.form['data_entrega']
        responsavel = request.form.get('responsavel')
        observacoes = request.form.get('observacoes')

        cur.execute("""
            INSERT INTO suprimentos_entregas (unidade_id, data_entrega, responsavel, observacoes)
            VALUES (%s, %s, %s, %s) RETURNING id
        """, (unidade_id, data_entrega, responsavel, observacoes))
        entrega_id = cur.fetchone()['id']

        tipos = request.form.getlist('tipo_suprimento[]')
        modelos = request.form.getlist('modelo_impressora[]')
        quantidades = request.form.getlist('quantidade[]')
        motivos_padrao = request.form.getlist('motivo_padrao[]')
        defeitos = request.form.getlist('defeito[]')
        motivos_outros = request.form.getlist('motivo[]')

        itens = []
        for tipo, modelo, qtd, mp, defeito, outro in zip(tipos, modelos, quantidades, motivos_padrao, defeitos, motivos_outros):
            tipo = (tipo or '').strip()
            if tipo:
                itens.append({
                    'tipo_suprimento': tipo,
                    'modelo_impressora': modelo,
                    'quantidade': int(qtd or 1),
                    'motivo_padrao': mp,
                    'defeito': defeito,
                    'outro': outro
                })

        if itens:
            faltantes = debitar_estoque_entrega(cur, entrega_id, itens, responsavel=responsavel)
            if faltantes:
                db.rollback()
                msgs = []
                for f in faltantes:
                    msgs.append(f"{f['tipo']} {f['modelo']}: solicitado {f['solicitado']}, em estoque {f['saldo']}")
                flash('Estoque insuficiente: ' + '; '.join(msgs), 'danger')
                cur.execute("""
                    SELECT emp.id as empresa_id, emp.nome as empresa_nome, emp.tipo as empresa_tipo,
                           u.id as unidade_id, u.nome as unidade_nome, u.setor
                    FROM empresas emp
                    LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
                    WHERE emp.ativo=1
                    ORDER BY emp.tipo DESC, emp.nome, u.nome
                """)
                locais = cur.fetchall()
                return render_template('suprimento_form.html', locais=locais, hoje=data_entrega)

        for item in itens:
            mp = item['motivo_padrao']
            defeito = item['defeito']
            outro = item['outro']
            motivo_texto = None
            if mp == 'outro':
                motivo_texto = (outro or '').strip() or None
            elif mp:
                motivo_texto = mp
            cur.execute("""
                INSERT INTO suprimentos_itens (entrega_id, tipo_suprimento, modelo_impressora, quantidade, motivo_padrao, defeito, motivo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (entrega_id, item['tipo_suprimento'], item['modelo_impressora'].strip() or None, item['quantidade'], mp or None, defeito.strip() or None, motivo_texto))

        db.commit()
        flash('Entrega de suprimentos registrada com sucesso!', 'success')
        return redirect(url_for('lista_suprimentos'))

    cur.execute("""
        SELECT emp.id as empresa_id, emp.nome as empresa_nome, emp.tipo as empresa_tipo,
               u.id as unidade_id, u.nome as unidade_nome, u.setor
        FROM empresas emp
        LEFT JOIN unidades u ON u.empresa_id = emp.id AND u.ativo=1
        WHERE emp.ativo=1
        ORDER BY emp.tipo DESC, emp.nome, u.nome
    """)
    locais = cur.fetchall()

    hoje = datetime.now().strftime('%Y-%m-%d')
    return render_template('suprimento_form.html', locais=locais, hoje=hoje)

@app.route('/suprimentos/excluir/<int:entrega_id>', methods=['POST'])
@login_required
def excluir_suprimento(entrega_id):
    db = get_db()
    cur = db.cursor()
    estornar_estoque_entrega(cur, entrega_id, responsavel=session.get('usuario'))
    cur.execute("DELETE FROM suprimentos_entregas WHERE id=%s", (entrega_id,))
    db.commit()
    flash('Registro excluido e estoque estornado com sucesso!', 'success')
    return redirect(url_for('lista_suprimentos'))


@app.route('/relatorio/mensal')
@login_required
def relatorio_mensal():
    db = get_db()
    cur = db.cursor()

    mes = request.args.get('mes', datetime.now().strftime('%Y-%m'))
    try:
        ano, numero_mes = mes.split('-')
        ano = int(ano)
        numero_mes = int(numero_mes)
    except Exception:
        ano = datetime.now().year
        numero_mes = datetime.now().month
        mes = f'{ano:04d}-{numero_mes:02d}'

    data_inicio = f'{ano}-{numero_mes:02d}-01'
    if numero_mes == 12:
        data_fim = f'{ano + 1}-01-01'
    else:
        data_fim = f'{ano}-{numero_mes + 1:02d}-01'

    # Resumo por tipo de suprimento
    cur.execute("""
        SELECT si.tipo_suprimento, si.modelo_impressora, SUM(si.quantidade) as total
        FROM suprimentos_itens si
        JOIN suprimentos_entregas se ON se.id = si.entrega_id
        WHERE se.data_entrega >= %s AND se.data_entrega < %s
        GROUP BY si.tipo_suprimento, si.modelo_impressora
        ORDER BY si.modelo_impressora, si.tipo_suprimento
    """, (data_inicio, data_fim))
    resumo_tipo = cur.fetchall()

    # Resumo por unidade
    cur.execute("""
        SELECT u.nome as unidade_nome, emp.nome as empresa_nome,
               si.tipo_suprimento, si.modelo_impressora, SUM(si.quantidade) as total
        FROM suprimentos_itens si
        JOIN suprimentos_entregas se ON se.id = si.entrega_id
        JOIN unidades u ON u.id = se.unidade_id
        JOIN empresas emp ON emp.id = u.empresa_id
        WHERE se.data_entrega >= %s AND se.data_entrega < %s
        GROUP BY u.nome, emp.nome, si.tipo_suprimento, si.modelo_impressora
        ORDER BY emp.nome, u.nome, si.modelo_impressora, si.tipo_suprimento
    """, (data_inicio, data_fim))
    resumo_unidade = cur.fetchall()

    # Lista de entregas detalhadas
    cur.execute("""
        SELECT se.id, se.data_entrega, se.data_registro, se.responsavel, se.observacoes,
               u.nome as unidade_nome, emp.nome as empresa_nome
        FROM suprimentos_entregas se
        JOIN unidades u ON u.id = se.unidade_id
        JOIN empresas emp ON emp.id = u.empresa_id
        WHERE se.data_entrega >= %s AND se.data_entrega < %s
        ORDER BY se.data_entrega DESC, se.data_registro DESC
    """, (data_inicio, data_fim))
    entregas = cur.fetchall()

    resultado_entregas = []
    for e in entregas:
        cur.execute("""
            SELECT tipo_suprimento, modelo_impressora, quantidade, motivo_padrao, defeito, motivo
            FROM suprimentos_itens WHERE entrega_id=%s
        """, (e['id'],))
        itens = cur.fetchall()
        itens_fmt = []
        for item in itens:
            nome = item['tipo_suprimento']
            if item['modelo_impressora']:
                nome += ' ' + item['modelo_impressora']
            motivo = item['motivo'] or item['motivo_padrao'] or ''
            itens_fmt.append({
                'nome': nome,
                'quantidade': item['quantidade'],
                'motivo': motivo
            })
        resultado_entregas.append({
            'id': e['id'],
            'data': e['data_entrega'].strftime('%d/%m/%Y') if hasattr(e['data_entrega'], 'strftime') else (str(e['data_entrega'])[:10] if e['data_entrega'] else '-'),
            'hora': e['data_registro'].strftime('%H:%M') if hasattr(e['data_registro'], 'strftime') else (str(e['data_registro'])[11:16] if e['data_registro'] else '-'),
            'unidade': e['unidade_nome'],
            'empresa': e['empresa_nome'],
            'responsavel': e['responsavel'],
            'observacoes': e['observacoes'],
            'itens': itens_fmt
        })

    # Agrupar por modelo para o resumo geral
    resumo_por_modelo = {}
    for r in resumo_tipo:
        modelo = r['modelo_impressora'] or 'Sem modelo'
        resumo_por_modelo.setdefault(modelo, []).append(r)

    # Agrupar por unidade
    unidades_map = {}
    for r in resumo_unidade:
        chave = (r['empresa_nome'], r['unidade_nome'])
        unidades_map.setdefault(chave, []).append(r)

    meses = [
        '2026-01','2026-02','2026-03','2026-04','2026-05','2026-06',
        '2026-07','2026-08','2026-09','2026-10','2026-11','2026-12'
    ]

    return render_template('relatorio_mensal.html',
        mes=mes, meses=meses, ano=ano, numero_mes=numero_mes,
        resumo_por_modelo=resumo_por_modelo,
        unidades_map=unidades_map,
        entregas=resultado_entregas)


@app.route('/estoque', methods=['GET'])
@login_required
def controle_estoque():
    db = get_db()
    cur = db.cursor()

    busca = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()

    sql = "SELECT * FROM estoque WHERE 1=1"
    params = []
    if busca:
        sql += """ AND (
            unaccent(lower(tipo_suprimento)) LIKE unaccent(lower(%s))
            OR unaccent(lower(modelo_impressora)) LIKE unaccent(lower(%s))
        )"""
        params.extend([f'%{busca}%', f'%{busca}%'])
    sql += " ORDER BY modelo_impressora, tipo_suprimento"
    cur.execute(sql, params)
    itens = cur.fetchall()

    resultado = []
    for item in itens:
        qtd = item['quantidade']
        minimo = item['estoque_minimo']
        if qtd == 0:
            st = 'zerado'
        elif qtd <= minimo:
            st = 'baixo'
        else:
            st = 'ok'
        if status and st != status:
            continue
        resultado.append({**item, 'status': st})

    # Agrupar por modelo
    grupos = {}
    for r in resultado:
        modelo = r['modelo_impressora'] or 'Sem modelo'
        grupos.setdefault(modelo, []).append(r)

    resumo = {'ok': 0, 'baixo': 0, 'zerado': 0}
    for r in resultado:
        resumo[r['status']] += 1

    return render_template('estoque.html', itens=resultado, grupos=grupos, busca=busca, filtro_status=status, resumo=resumo)


@app.route('/estoque/entrada', methods=['GET', 'POST'])
@login_required
def estoque_entrada():
    db = get_db()
    cur = db.cursor()

    if request.method == 'POST':
        tipo = request.form.get('tipo_suprimento', '').strip()
        cor = request.form.get('cor_selecionada', '').strip()
        modelo = request.form.get('modelo_impressora', '').strip().upper()
        quantidade = int(request.form.get('quantidade') or 0)
        estoque_minimo = int(request.form.get('estoque_minimo') or 1)
        motivo = request.form.get('motivo', '').strip() or None
        responsavel = request.form.get('responsavel', '').strip() or session.get('usuario')

        if not tipo or quantidade <= 0:
            flash('Informe o tipo e a quantidade.', 'danger')
            return redirect(url_for('estoque_entrada'))

        tipo_final = f"{tipo} {cor}".strip() if cor else tipo
        estoque_id, saldo = buscar_ou_criar_estoque(cur, tipo_final, modelo)

        cur.execute(
            "UPDATE estoque SET estoque_minimo=%s WHERE id=%s",
            (estoque_minimo, estoque_id)
        )
        movimentar_estoque(
            cur, estoque_id, 'entrada', quantidade, saldo,
            motivo=motivo or 'Entrada manual de estoque',
            responsavel=responsavel
        )
        db.commit()
        flash(f'Entrada registrada: {tipo_final} {modelo} (+{quantidade})', 'success')
        return redirect(url_for('controle_estoque'))

    return render_template('estoque_entrada.html')


@app.route('/estoque/ajuste/<int:estoque_id>', methods=['GET', 'POST'])
@login_required
def estoque_ajuste(estoque_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM estoque WHERE id=%s", (estoque_id,))
    item = cur.fetchone()
    if not item:
        flash('Item nao encontrado.', 'danger')
        return redirect(url_for('controle_estoque'))

    if request.method == 'POST':
        nova_qtd = int(request.form.get('quantidade') or 0)
        estoque_minimo = int(request.form.get('estoque_minimo') or item['estoque_minimo'])
        motivo = request.form.get('motivo', '').strip() or 'Ajuste manual'
        responsavel = request.form.get('responsavel', '').strip() or session.get('usuario')

        if nova_qtd < 0:
            flash('Quantidade nao pode ser negativa.', 'danger')
            return redirect(url_for('estoque_ajuste', estoque_id=estoque_id))

        diferenca = nova_qtd - item['quantidade']
        tipo_movimento = 'ajuste'
        if diferenca > 0:
            tipo_movimento = 'entrada'
        elif diferenca < 0:
            tipo_movimento = 'saida'

        cur.execute("UPDATE estoque SET estoque_minimo=%s WHERE id=%s", (estoque_minimo, estoque_id))
        if diferenca != 0:
            movimentar_estoque(
                cur, estoque_id, tipo_movimento, diferenca, item['quantidade'],
                motivo=motivo, responsavel=responsavel
            )
        db.commit()
        flash('Ajuste salvo com sucesso!', 'success')
        return redirect(url_for('controle_estoque'))

    return render_template('estoque_ajuste.html', item=item)


@app.route('/estoque/historico/<int:estoque_id>')
@login_required
def estoque_historico(estoque_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM estoque WHERE id=%s", (estoque_id,))
    item = cur.fetchone()
    if not item:
        flash('Item nao encontrado.', 'danger')
        return redirect(url_for('controle_estoque'))

    cur.execute("""
        SELECT * FROM estoque_movimentacoes
        WHERE estoque_id=%s
        ORDER BY data_movimento DESC, id DESC
    """, (estoque_id,))
    movimentacoes = cur.fetchall()

    return render_template('estoque_historico.html', item=item, movimentacoes=movimentacoes)


@app.route('/api/estoque/historico')
@login_required
def api_estoque_historico():
    tipo = request.args.get('tipo', '').strip()
    modelo = request.args.get('modelo', '').strip().upper()
    if not tipo:
        return jsonify({'erro': 'Informe o tipo.'}), 400
    db = get_db()
    cur = db.cursor()
    estoque_id, _ = buscar_ou_criar_estoque(cur, tipo, modelo)
    cur.execute("""
        SELECT id, tipo_movimento, quantidade, saldo_antes, saldo_depois, motivo, responsavel, data_movimento
        FROM estoque_movimentacoes
        WHERE estoque_id=%s
        ORDER BY data_movimento DESC, id DESC
        LIMIT 100
    """, (estoque_id,))
    rows = cur.fetchall()
    movimentacoes = []
    for r in rows:
        movimentacoes.append({
            'id': r['id'],
            'tipo_movimento': r['tipo_movimento'],
            'quantidade': r['quantidade'],
            'saldo_antes': r['saldo_antes'],
            'saldo_depois': r['saldo_depois'],
            'motivo': r['motivo'],
            'responsavel': r['responsavel'],
            'data_movimento': r['data_movimento'].strftime('%d/%m/%Y %H:%M') if r['data_movimento'] else None
        })
    return jsonify({'movimentacoes': movimentacoes})


@app.route('/api/suprimentos/historico')
@login_required
def api_suprimentos_historico():
    db = get_db()
    cur = db.cursor()

    limite = request.args.get('limite', 50, type=int)
    unidade_id = request.args.get('unidade_id', type=int)

    params = []
    sql = """
        SELECT se.id, se.data_entrega, se.data_registro, se.responsavel, se.observacoes,
               u.nome as unidade_nome, emp.nome as empresa_nome
        FROM suprimentos_entregas se
        JOIN unidades u ON u.id = se.unidade_id
        JOIN empresas emp ON emp.id = u.empresa_id
        WHERE 1=1
    """
    if unidade_id:
        sql += " AND se.unidade_id=%s"
        params.append(unidade_id)
    sql += " ORDER BY se.data_registro DESC LIMIT %s"
    params.append(limite)

    cur.execute(sql, params)
    entregas = cur.fetchall()

    resultado = []
    for e in entregas:
        cur.execute("""
            SELECT tipo_suprimento, modelo_impressora, quantidade, motivo_padrao, defeito, motivo
            FROM suprimentos_itens
            WHERE entrega_id=%s
        """, (e['id'],))
        itens = cur.fetchall()
        itens_fmt = []
        for item in itens:
            nome = item['tipo_suprimento']
            if item['modelo_impressora']:
                nome += ' ' + item['modelo_impressora']
            motivo = item['motivo'] or item['motivo_padrao'] or ''
            itens_fmt.append({
                'nome': nome,
                'quantidade': item['quantidade'],
                'motivo': motivo
            })
        resultado.append({
            'id': e['id'],
            'data': e['data_entrega'] or '-',
            'hora': e['data_registro'].strftime('%H:%M') if e['data_registro'] else '-',
            'unidade': e['unidade_nome'],
            'empresa': e['empresa_nome'],
            'responsavel': e['responsavel'],
            'observacoes': e['observacoes'],
            'itens': itens_fmt
        })

    return jsonify({'entregas': resultado})


@app.route('/api/estoque/ajuste', methods=['POST'])
@login_required
def api_estoque_ajuste():
    tipo = request.form.get('tipo_suprimento', '').strip()
    cor = request.form.get('cor_selecionada', '').strip()
    modelo = request.form.get('modelo_impressora', '').strip().upper()
    nova_qtd = int(request.form.get('quantidade') or 0)
    motivo = request.form.get('motivo', '').strip() or 'Ajuste via app mobile'
    responsavel = request.form.get('responsavel', '').strip() or session.get('usuario')

    if not tipo or nova_qtd < 0:
        return jsonify({'erro': 'Informe o tipo e uma quantidade válida.'}), 400

    tipo_final = f"{tipo} {cor}".strip() if cor else tipo
    db = get_db()
    cur = db.cursor()
    estoque_id, saldo = buscar_ou_criar_estoque(cur, tipo_final, modelo)

    diferenca = nova_qtd - saldo
    tipo_movimento = 'ajuste'
    if diferenca > 0:
        tipo_movimento = 'entrada'
    elif diferenca < 0:
        tipo_movimento = 'saida'

    if diferenca != 0:
        movimentar_estoque(
            cur, estoque_id, tipo_movimento, diferenca, saldo,
            motivo=motivo, responsavel=responsavel
        )
    db.commit()
    return jsonify({'ok': True, 'mensagem': f'{tipo_final} {modelo}: saldo ajustado para {nova_qtd}'})


@app.route('/api/estoque/saldo')
@login_required
def api_estoque_saldo():
    tipo = request.args.get('tipo', '').strip()
    modelo = request.args.get('modelo', '').strip().upper()
    if not tipo:
        return jsonify({'saldo': 0})
    db = get_db()
    cur = db.cursor()
    _, saldo = buscar_ou_criar_estoque(cur, tipo, modelo)
    return jsonify({'saldo': saldo})


@app.route('/api/estoque/entrada', methods=['POST'])
@login_required
def api_estoque_entrada():
    tipo = request.form.get('tipo_suprimento', '').strip()
    cor = request.form.get('cor_selecionada', '').strip()
    modelo = request.form.get('modelo_impressora', '').strip().upper()
    quantidade = int(request.form.get('quantidade') or 0)
    estoque_minimo = int(request.form.get('estoque_minimo') or 1)
    motivo = request.form.get('motivo', '').strip() or None
    responsavel = request.form.get('responsavel', '').strip() or session.get('usuario')

    if not tipo or quantidade <= 0:
        return jsonify({'erro': 'Informe o tipo e a quantidade.'}), 400

    tipo_final = f"{tipo} {cor}".strip() if cor else tipo
    db = get_db()
    cur = db.cursor()
    estoque_id, saldo = buscar_ou_criar_estoque(cur, tipo_final, modelo)
    cur.execute("UPDATE estoque SET estoque_minimo=%s WHERE id=%s", (estoque_minimo, estoque_id))
    movimentar_estoque(
        cur, estoque_id, 'entrada', quantidade, saldo,
        motivo=motivo or 'Entrada via app mobile',
        responsavel=responsavel
    )
    db.commit()
    return jsonify({'ok': True, 'mensagem': f'{tipo_final} {modelo}: +{quantidade}'})


# Para produção (Render / Gunicorn)
# Inicializa o banco automaticamente se estiver vazio
with app.app_context():
    init_db()
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) as count FROM equipamentos")
    if cur.fetchone()['count'] == 0:
        try:
            importar_planilha_para_banco()
        except Exception as e:
            print(f"Aviso: não foi possível importar planilha automaticamente: {e}")

if __name__ == '__main__':
    print("Acesse: http://localhost:5000")
    app.run(debug=True, port=5000)
