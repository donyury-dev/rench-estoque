import pandas as pd
import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\Kaio\.verdent\verdent-projects\agora-como-um-novo\data_atual.xlsx',
    data_only=True
)

lines = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"===== {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) =====")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        if any(v is not None for v in row):
            lines.append(str(row))
    lines.append("")

with open(r'C:\Users\Kaio\.verdent\verdent-projects\agora-como-um-novo\temp_planilha.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('ok')
